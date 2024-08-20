'''
用于配置excel的写入格式内容：
    把断言的结果pass和Failed更突出一些
    PatternFill类用于定义颜色
    Font类用于定于格式
'''

from openpyxl.styles import PatternFill, Font


#pass写入配置
def pass_(cell,row,column):
    cell(row=row,column=column).value='Pass'
    #单元格显示，绿色加粗
    cell(row=row, column=column).fill= PatternFill('solid',fgColor='AACF91')
    cell(row=row, column=column).font= Font(bold=True)

#Failed写入配置
def failed_(cell,row,column):
    cell(row=row,column=column).value='Failed'
    #单元格显示，红色色加粗
    cell(row=row, column=column).fill= PatternFill('solid',fgColor='FF0000')
    cell(row=row, column=column).font= Font(bold=True)
