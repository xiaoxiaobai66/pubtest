'''
excel 文件读取类，用于实现测试用例文件的读取与执行

'''
import pathlib
import openpyxl
from KeyWord_Excel.key_log.logmoudle import test_log2

#解析测试用例中测试参数单元格的内容，并转换为字典的形态
# from web_automation.moudle.web_key import WebUiKeys
from web_automation.moudle import excel_cof
from KeyWord_Excel.keyword_packge.web_key import WebUiKeys


log2=test_log2()# loggeer组件封装实例化

def dealtodict(value):
    # excel文件：参数的处理--通过一个dict来接收所有的参数内容。便于定值不定长的参数形态
    # 参数最终形态:'type_=Chrome' 改变为 {tpye_:'Chrome'}
    dict1 = dict()
    if value:  # 如果elvalue[4]为空，后续切分时会报错，所以这里先过滤
        everyp = value.split(",")  # 先用逗号 切分有多少个入参
        for onebyone in everyp:  # 遍历每个入参
            one = onebyone.split("=", 1)  # 每个入参再切分成key和value的关系。
            # 由于xpath中的内容也会被切分掉，所以限制只切一次
            dict1[one[0]] = one[1]
        # print(dict1)
        # print(everyp)
    else:
        # dict1=None
        pass  #如果为空则不处理
    return dict1

def excel_read(filename,replay=False,replay_case=None):
    # #1、文件路径 2、由main文件调用时，传入文件参数--filename
    # file= pathlib.Path(__file__).resolve().parents[1]/f'数据驱动/{filename}'
    #获取excel的内容
    data=openpyxl.load_workbook(filename) #"./数据驱动/testui.xlsx"
    log2.info(f'***********开始执行{filename}文件用例***************')
    # sheet=data['Sheet1']
    # 记录成功失败
    sucesscase = 0
    failcase = 0
    failcount = dict()
    failcasename = list()
    if replay:
        all_sheet=replay_case
    else:
        all_sheet=data.sheetnames
    #遍历所有sheet页
    for oneofsheet in all_sheet:
        sheet=data[oneofsheet]
        if replay:
            log2.info(f'***********{oneofsheet}失败用例重试***************')
        else:
            log2.info(f'***********开始执行{oneofsheet}用例***************')
        # print(f'***********开始执行{oneofsheet}用例***************')
        for elvalue in sheet.values: #遍历表中每行数据，以元组形式返回
            # #获取测试用例的正文内容
            if type(elvalue[0]) is int: #遍历每行的第一位，直到数字才开始输出
                # print(elvalue)
                #用例具体描述可以做用例日志的输出
                # print(f'**********正在执行：{elvalue[2]}')
                log2.info(f'**********正在执行：{elvalue[2]}')
                #excel文件：参数的处理--通过一个dict来接收所有的参数内容。便于定值不定长的参数形态
                #参数最终形态:'type_=Chrome' 改变为 {tpye_:'Chrome'}
                # dict1=dict()
                # if elvalue[4]: #如果elvalue[4]为空，后续切分时会报错，所以这里先过滤
                #     everyp=elvalue[4].split(",") #先用逗号 切分有多少个入参
                #     for onebyone in everyp: #遍历每个入参
                #         one=onebyone.split("=",1) # 每个入参再切分成key和value的关系。
                #                                   # 由于xpath中的内容也会被切分掉，所以限制只切一次
                #         dict1[one[0]]=one[1]
                #     # print(dict1)
                #     # print(everyp)
                # else:
                #     dict1=None
                # print(dict1)
                testdata_dict=dealtodict(elvalue[4])
                # print(dealtodict(elvalue[4]))
                '''
                    #excel文件：调用的函数是elvalue[1]
                        操作行为的调用分为以下几种不同的类型：
                        1.实例化 driver
                        2.基于实例化对象进行的操作行为
                        3.断言机制：有预期与实际的对比，以及有单元格测试结果的写入
                '''
                '''
                    对反射的解析：
                    excel用例中第一行：open_browser
                    第二行open:getattr(driver,'open')(**testdata_dict)
                            driver.open(**testdata_dict)
                            driver.open(url="https://litemall.hogwarts.ceshiren.com/vue/index.html#/")
                    第三行click： getattr(driver,'click')(**testdata_dict) 
                            driver.click(**testdata_dict)
                            driver.input(by='xpath',value="//i[@class='van-icon van-icon-wode']")   
                    第四行input: getattr(driver,'input')(**testdata_dict)
                            driver.input(**testdata_dict)
                            driver.input(by='xpath',value="//input[@name='user']",txt='user123')                    
                        
                '''
                #实例化操作
                # print(elvalue[1])
                if elvalue[1]=='open_browser':
                    driver=WebUiKeys(**testdata_dict) #**testdata_dict为解包操作,即将{'type_': 'Chrome'}解为type_='Chrome'传入

                #断言行为：基于断言的返回结果来判断测试的成功失败，并进行写入操作
                elif 'assert' in elvalue[1]:
                    # status=getattr(driver,elvalue[1])(**testdata_dict) #两种都可以，使用下面的话需要将expected写在excel的预期栏中
                    status=getattr(driver, elvalue[1])(expected=elvalue[5], **testdata_dict)

                    #基于status写入测试结果
                    if status:
                        # sheet.cell(row=elvalue[0]+2,column=7).value='Pass'
                        excel_cof.pass_(sheet.cell,row=elvalue[0]+2,column=7)
                        #统计成功用例
                        sucesscase+=1
                    else:
                        # sheet.cell(row=elvalue[0] + 2, column=7).value = 'Failed'
                        excel_cof.failed_(sheet.cell, row=elvalue[0] + 2, column=7)
                        #统计失败用例，并且将失败的用例sheet名放进例表中
                        failcase+=1
                        failcasename.append(oneofsheet)

                    #保存excel，写入即保存，防止后续代码报错中断，导致之前的测试结果都没有了
                    data.save(filename)  #"./数据驱动/testui.xlsx"
                #常规操作
                else:
                    # if testdata_dict: #解决拆包为None时报错方法一：将dealtodict函数中的else:改为pass；方法二：启用本行if、else
                    getattr(driver,elvalue[1])(**testdata_dict)
                    # else:
                    #     getattr(driver, elvalue[1])() #为空时没有解包操作
    failcount[filename]=failcasename
    print(failcount)
    #判断是否是失败重试流程
    if replay==False:
        log2.info(f'*********{filename}文件执行完毕：成功{sucesscase}条，失败{failcase}条  失败用例：{failcount[filename]}***********')
    else:
        log2.info(f'{filename}文件失败用例重试：成功{sucesscase}条，失败{failcase}条  失败用例：{failcount[filename]}***********')
        #无论重试结果如何，只重试一次
        failcount[filename]=[]
    return failcount
if __name__ == '__main__':
    excel_read()