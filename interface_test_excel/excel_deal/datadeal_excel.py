#读取excel数据

import openpyxl

class exceData:
    def getExcel(self,file):
        #找到需要操作的excel  '../data/test_account.xlsx'
        workb=openpyxl.load_workbook(file)
        #找到sheet页
        sheet1=workb['login']
        #拿到sheet页的数据, 迭代对象sheet1.iter_rows()
        row_sheet=sheet1.iter_rows()
        #遍历可迭代数据
        listall=[]
        for rowdata in row_sheet:
            # print(rowdata) #是一个元组对象，可通过下标+value提取
            if type(rowdata[0].value) is int:
                # print(rowdata)

                #为方便后续操作，将元组中的值，放到列表中
                #遍历行数据中的 每一列
                listrow=[]
                for col in rowdata:
                    listrow.append(col.value)
                # print(listrow) #有5条数据，表示5个用例，将这些用例再放到一个总列表中,只要拿到这个总列表，相当于拿到这个sheet中所有用例
                listall.append(listrow)#将这些用例再放到一个总列表中,只要拿到这个总列表，相当于拿到这个sheet中所有用例

        # print(listall)
        return listall

if __name__ == '__main__':
    excel1=exceData().getExcel('../data/test_account.xlsx')
    print(excel1)